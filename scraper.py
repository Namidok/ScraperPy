# scraper.py - Multi-Platform Job Scraper with Separate Excel Sheets (Final Version)

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from datetime import datetime
import time
import random
import os

EXCEL_FILE = "jobs.xlsx"

def setup_driver():
    """
    Configure Chrome driver with anti-detection settings
    """
    chrome_options = Options()
    
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_argument('--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    
    return driver


def create_excel_if_not_exists():
    """
    Create Excel file with separate sheets for StepStone and LinkedIn ONLY
    """
    if os.path.exists(EXCEL_FILE):
        # Remove old 'Jobs' sheet if it exists
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            if 'Jobs' in wb.sheetnames:
                del wb['Jobs']
                wb.save(EXCEL_FILE)
                print("✅ Removed old 'Jobs' sheet")
        except:
            pass
        return
    
    print(f"📄 Creating new Excel file: {EXCEL_FILE}")
    
    # Create Excel with ONLY StepStone and LinkedIn sheets
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        for sheet_name in ['StepStone', 'LinkedIn']:
            data = {
                'Job Title': [],
                'Company': [],
                'Location': [],
                'Description': [],
                'URL': [],
                'Posted Date': [],
                'Date Added': [],
                'Applied': [],
                'Application Date': [],
                'Status': [],
                'Notes': []
            }
            
            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Format the Excel file
    wb = openpyxl.load_workbook(EXCEL_FILE)
    
    column_widths = {
        'A': 25,  # Job Title
        'B': 20,  # Company
        'C': 15,  # Location
        'D': 30,  # Description
        'E': 50,  # URL
        'F': 12,  # Posted Date
        'G': 12,  # Date Added
        'H': 8,   # Applied
        'I': 15,  # Application Date
        'J': 12,  # Status
        'K': 20   # Notes
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        ws.freeze_panes = 'A2'
    
    wb.save(EXCEL_FILE)
    print(f"✅ Excel file created with StepStone and LinkedIn sheets only\n")


def read_existing_urls(sheet_name):
    """
    Read existing job URLs from a specific sheet to avoid duplicates
    """
    if not os.path.exists(EXCEL_FILE):
        return set()
    
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
        return set(df['URL'].dropna())
    except Exception as e:
        print(f"⚠️  Error reading existing URLs from {sheet_name}: {e}")
        return set()


def append_jobs_to_excel(new_jobs, platform):
    """
    Append new jobs to the corresponding Excel sheet
    """
    if not new_jobs:
        print(f"⚠️  No new jobs to add for {platform}")
        return
    
    sheet_name = platform
    
    # Read existing data from the specific sheet
    try:
        df_existing = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
    except:
        df_existing = pd.DataFrame()
    
    # Create dataframe from new jobs
    today = datetime.now().strftime("%Y-%m-%d")
    new_data = []
    
    for job in new_jobs:
        new_data.append({
            'Job Title': job['title'],
            'Company': job['company'],
            'Location': job['location'],
            'Description': job['description'],
            'URL': job['url'],
            'Posted Date': job['posted_date'],
            'Date Added': today,
            'Applied': 'No',
            'Application Date': '',
            'Status': 'New',
            'Notes': ''
        })
    
    df_new = pd.DataFrame(new_data)
    
    # Remove duplicates (if job URL already exists, skip it)
    if not df_existing.empty:
        existing_urls = set(df_existing['URL'].dropna())
        df_new = df_new[~df_new['URL'].isin(existing_urls)]
    
    if df_new.empty:
        print(f"✅ All {platform} jobs are duplicates - no new jobs added")
        return
    
    # Append to existing data
    if df_existing.empty:
        df_combined = df_new
    else:
        df_combined = pd.concat([df_existing, df_new], ignore_index=True)
    
    # Write back to Excel - need to rewrite entire workbook
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_combined.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Re-apply formatting
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    
    column_widths = {
        'A': 25,
        'B': 20,
        'C': 15,
        'D': 30,
        'E': 50,
        'F': 12,
        'G': 12,
        'H': 8,
        'I': 15,
        'J': 12,
        'K': 20
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = 'A2'
    wb.save(EXCEL_FILE)
    
    print(f"✅ Added {len(df_new)} new {platform} jobs to {EXCEL_FILE}")
    print(f"📊 Total {platform} jobs in Excel: {len(df_combined)}")


def scrape_stepstone(driver, search_term="werkstudent IT", location="Berlin", max_pages=2):
    """
    Scrape StepStone for job listings with English and 24-hour filters
    Extracts DIRECT job URLs by clicking on each job
    """
    jobs = []
    print(f"\n🔍 [STEPSTONE] Searching: {search_term} in {location}")
    
    try:
        # Build initial URL
        search_slug = search_term.lower().replace(' ', '-')
        url = f"https://www.stepstone.de/jobs/{search_slug}/in-{location.lower()}"
        
        print(f"  🌐 Loading StepStone: {url}")
        driver.get(url)
        time.sleep(random.uniform(4, 6))
        
        # Accept cookies if present
        try:
            cookie_button = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Alle akzeptieren') or contains(text(), 'Accept')]"))
            )
            cookie_button.click()
            print("  🍪 Accepted cookies")
            time.sleep(2)
        except:
            print("  ℹ️  No cookie banner found")
        
        # Apply 24-hour filter
        try:
            print("  ⏰ Applying 24-hour filter...")
            time_filter = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), '24') or contains(@aria-label, '24 Stunden')]"))
            )
            time_filter.click()
            time.sleep(2)
            print("  ✅ 24-hour filter applied")
        except Exception as e:
            print(f"  ⚠️  Could not apply 24-hour filter: {e}")
        
        # Apply English language filter
        try:
            print("  🌍 Applying English language filter...")
            
            try:
                filter_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Filter') or contains(text(), 'Alle Filter')]")
                filter_button.click()
                time.sleep(2)
            except:
                pass
            
            english_checkbox = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//label[contains(text(), 'Englisch') or contains(text(), 'English')]"))
            )
            english_checkbox.click()
            time.sleep(2)
            print("  ✅ English language filter applied")
            
            try:
                apply_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Anwenden') or contains(text(), 'Apply')]")
                apply_button.click()
                time.sleep(3)
            except:
                pass
            
        except Exception as e:
            print(f"  ⚠️  Could not apply English filter: {e}")
        
        # Scrape pages
        for page in range(1, max_pages + 1):
            try:
                if page > 1:
                    current_url = driver.current_url
                    if '?' in current_url:
                        next_url = f"{current_url}&page={page}"
                    else:
                        next_url = f"{current_url}?page={page}"
                    driver.get(next_url)
                    time.sleep(random.uniform(3, 5))
                
                print(f"  📄 Scraping page {page}...")
                
                try:
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.TAG_NAME, "article"))
                    )
                except:
                    print(f"  ⚠️  Timeout waiting for job cards")
                    break
                
                # Get job card elements with Selenium to extract proper URLs
                job_card_elements = driver.find_elements(By.TAG_NAME, "article")
                
                if not job_card_elements:
                    print(f"  ⚠️  No jobs found on page {page}")
                    break
                
                print(f"  ✅ Found {len(job_card_elements)} jobs")
                
                for card in job_card_elements:
                    try:
                        # Get the title link
                        title_link = card.find_element(By.XPATH, ".//a[contains(@href, '/stellenangebote')]")
                        job_title = title_link.text.strip()
                        job_url = title_link.get_attribute('href')
                        
                        if not job_url:
                            continue
                        
                        # Make sure URL is complete
                        if not job_url.startswith('http'):
                            job_url = 'https://www.stepstone.de' + job_url
                        
                        # Parse with BeautifulSoup for other fields
                        soup = BeautifulSoup(str(card.get_attribute('outerHTML')), 'html.parser')
                        
                        # Company
                        company_elem = soup.find('span', class_=lambda x: x and 'company' in str(x).lower() if x else False)
                        if not company_elem:
                            company_elem = soup.find_all('span')
                            company_elem = company_elem[1] if len(company_elem) > 1 else None
                        company = company_elem.get_text(strip=True) if company_elem else "N/A"
                        
                        # Location
                        location_elem = soup.find('span', class_=lambda x: x and 'location' in str(x).lower() if x else False)
                        job_location = location_elem.get_text(strip=True) if location_elem else location
                        
                        # Description
                        desc_elem = soup.find('p')
                        description = desc_elem.get_text(strip=True) if desc_elem else ""
                        
                        # Posted date
                        date_elem = soup.find('time')
                        posted_date = date_elem.get('datetime', 'N/A') if date_elem else "N/A"
                        
                        if not any(j['url'] == job_url for j in jobs):
                            jobs.append({
                                'title': job_title,
                                'company': company,
                                'location': job_location,
                                'description': description[:500],
                                'url': job_url,
                                'posted_date': posted_date,
                                'platform': 'StepStone',
                                'search_term': search_term
                            })
                            print(f"    ➕ {job_title} @ {company}")
                    
                    except Exception as e:
                        continue
                
                time.sleep(random.uniform(2, 4))
            
            except Exception as e:
                print(f"  ❌ Error on page {page}: {e}")
                break
    
    except Exception as e:
        print(f"  ❌ Error in StepStone scraper: {e}")
    
    return jobs


def scrape_linkedin(driver, search_term="werkstudent IT", location="Berlin", max_pages=2):
    """
    Scrape LinkedIn for job listings with 24-hour filter
    """
    jobs = []
    print(f"\n🔍 [LINKEDIN] Searching: {search_term} in {location}")
    
    for page in range(max_pages):
        try:
            start = page * 25
            url = f"https://www.linkedin.com/jobs/search?keywords={search_term}&location={location}&f_TPR=r86400&sortBy=DD&start={start}"
            print(f"  📄 Page {page + 1}: Past 24 hours filter active")
            
            driver.get(url)
            time.sleep(random.uniform(4, 6))
            
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            
            job_cards = soup.find_all('div', class_=lambda x: x and 'base-card' in x if x else False)
            
            if not job_cards:
                job_cards = soup.find_all('li', class_=lambda x: x and 'job' in str(x).lower() if x else False)
            
            if not job_cards:
                print(f"  ⚠️  No jobs found")
                break
            
            print(f"  ✅ Found {len(job_cards)} jobs")
            
            for card in job_cards:
                try:
                    title_elem = card.find('h3', class_=lambda x: x and 'base-search-card__title' in x if x else False)
                    if not title_elem:
                        title_elem = card.find('a', href=lambda x: x and '/jobs/view/' in x if x else False)
                    
                    if not title_elem:
                        continue
                    
                    job_title = title_elem.get_text(strip=True)
                    
                    link_elem = card.find('a', href=lambda x: x and '/jobs/view/' in x if x else False)
                    job_url = link_elem['href'] if link_elem else None
                    if job_url and not job_url.startswith('http'):
                        job_url = 'https://www.linkedin.com' + job_url
                    
                    company_elem = card.find('h4', class_=lambda x: x and 'base-search-card__subtitle' in x if x else False)
                    if not company_elem:
                        company_elem = card.find('a', class_=lambda x: x and 'company' in str(x).lower() if x else False)
                    company = company_elem.get_text(strip=True) if company_elem else "N/A"
                    
                    location_elem = card.find('span', class_=lambda x: x and 'job-search-card__location' in x if x else False)
                    job_location = location_elem.get_text(strip=True) if location_elem else location
                    
                    date_elem = card.find('time')
                    posted_date = date_elem.get('datetime', 'N/A') if date_elem else "24h"
                    
                    if job_url and not any(j['url'] == job_url for j in jobs):
                        jobs.append({
                            'title': job_title,
                            'company': company,
                            'location': job_location,
                            'description': "",
                            'url': job_url,
                            'posted_date': posted_date,
                            'platform': 'LinkedIn',
                            'search_term': search_term
                        })
                        print(f"    ➕ {job_title} @ {company}")
                
                except Exception as e:
                    continue
            
            time.sleep(random.uniform(3, 5))
        
        except Exception as e:
            print(f"  ❌ Error on page {page + 1}: {e}")
            break
    
    return jobs


def scrape_all_platforms(search_terms=["werkstudent IT"], location="Berlin", max_pages=2):
    """
    Scrape StepStone and LinkedIn
    """
    print("🚀 Starting Multi-Platform Job Scraper (24-hour filter active)")
    print("=" * 80)
    
    stepstone_jobs = []
    linkedin_jobs = []
    driver = setup_driver()
    
    try:
        for search_term in search_terms:
            stepstone_jobs.extend(scrape_stepstone(driver, search_term, location, max_pages))
            linkedin_jobs.extend(scrape_linkedin(driver, search_term, location, max_pages))
    
    finally:
        driver.quit()
        print("\n🔒 Browser closed")
    
    return stepstone_jobs, linkedin_jobs


if __name__ == "__main__":
    print("=" * 80)
    print("WERKSTUDENT IT JOB APPLICATION TRACKER")
    print("=" * 80)
    
    # Create Excel file with separate sheets (StepStone and LinkedIn ONLY)
    create_excel_if_not_exists()
    
    search_keywords = [
        "werkstudent IT",
        "working student software"
    ]
    
    # Scrape jobs from all platforms
    stepstone_jobs, linkedin_jobs = scrape_all_platforms(search_terms=search_keywords, location="Berlin", max_pages=2)
    
    print("\n" + "=" * 80)
    print(f"✅ TOTAL JOBS SCRAPED: {len(stepstone_jobs) + len(linkedin_jobs)}")
    print("=" * 80)
    print(f"  StepStone: {len(stepstone_jobs)} jobs")
    print(f"  LinkedIn: {len(linkedin_jobs)} jobs")
    
    # Append to Excel (separate sheets)
    print("\n" + "=" * 80)
    print("UPDATING EXCEL FILE")
    print("=" * 80)
    
    append_jobs_to_excel(stepstone_jobs, 'StepStone')
    append_jobs_to_excel(linkedin_jobs, 'LinkedIn')
    
    # Show sample results
    print("\n" + "=" * 80)
    print("SAMPLE RESULTS (First 3 from each):")
    print("=" * 80)
    
    if stepstone_jobs:
        print("\n🏢 STEPSTONE:")
        for i, job in enumerate(stepstone_jobs[:3], 1):
            print(f"{i}. {job['title']}")
            print(f"   Company: {job['company']}")
            print(f"   URL: {job['url']}")
    
    if linkedin_jobs:
        print("\n💼 LINKEDIN:")
        for i, job in enumerate(linkedin_jobs[:3], 1):
            print(f"{i}. {job['title']}")
            print(f"   Company: {job['company']}")
            print(f"   URL: {job['url']}")
    
    print("\n" + "=" * 80)
    print(f"📁 Jobs saved to: {EXCEL_FILE}")
    print("✅ Only StepStone and LinkedIn sheets (no 'Jobs' sheet)")
    print("✅ Direct job URLs for StepStone (clickable & working)")
    print("=" * 80)
